"""
Router Compare — /api/generate/compare
========================================
Génération de rapports comparatifs multi-groupes (candidats / sociétés).

Endpoints :
  POST /generate/compare              → lance la comparaison
  GET  /generate/compare/stream/{id} → flux SSE de progression
  GET  /generate/compare/download/{id} → télécharge le fichier Excel
"""

import asyncio
import json
import re
import uuid
from datetime import datetime, timezone
from pathlib import Path

import aiofiles
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from config import get_settings
from database import get_db
from logger import get_logger
from models.document import Document
from models.job import Job
from models.template import Template
from services.ollama_service import OllamaService

log = get_logger(__name__)
settings = get_settings()
router = APIRouter()

# Cache en mémoire : job_id → état de la comparaison
_compare_cache: dict[str, dict] = {}
# Structure : {
#   "events": [{"groupe": str, "statut": str, "index": int, "total": int}],
#   "statut": "running" | "complete" | "failed",
#   "fichier": str | None,   # chemin absolu du fichier Excel généré
#   "erreur": str | None,
# }

MAX_CHARS_PAR_DOC = 20_000   # Tronquer les gros docs pour tenir dans le contexte


class GroupeRequest(BaseModel):
    nom: str = Field(..., min_length=1, description="Nom du candidat / société")
    document_ids: list[str] = Field(..., min_items=1)


class CompareRequest(BaseModel):
    groupes: list[GroupeRequest] = Field(..., min_items=2, description="Au moins 2 groupes")
    template_id: str = Field(..., description="ID du template Excel")
    model: str | None = Field(default=None)
    instructions: str | None = Field(default=None, description="Instructions supplémentaires pour le LLM")


def _lire_colonnes_template(chemin: Path) -> list[str]:
    """Lit les en-têtes de la première ligne du template Excel."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(chemin), read_only=True, data_only=True)
        ws = wb.active
        colonnes = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ()):
            if cell and isinstance(cell, str) and cell.strip():
                colonnes.append(cell.strip())
        wb.close()
        return colonnes
    except Exception as e:
        log.warning("Impossible de lire les colonnes du template", erreur=str(e))
        return []


def _remplir_excel(template_path: Path, groupes_data: list[dict], colonnes: list[str]) -> Path:
    """
    Remplit le template Excel avec les données extraites.
    groupes_data : [{nom: str, valeurs: {colonne: valeur}}]
    Retourne le chemin du fichier généré.
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(template_path))
    ws = wb.active

    # Trouver la ligne des en-têtes (ligne 1) et écrire les données à partir de la ligne 2
    for row_idx, groupe in enumerate(groupes_data, start=2):
        for col_idx, colonne in enumerate(colonnes, start=1):
            valeur = groupe.get("valeurs", {}).get(colonne, "N/A")
            ws.cell(row=row_idx, column=col_idx, value=valeur)

    exports_dir = Path(settings.storage_exports)
    exports_dir.mkdir(parents=True, exist_ok=True)
    nom_fichier = f"comparatif_{datetime.now(tz=timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    sortie = exports_dir / nom_fichier
    wb.save(str(sortie))
    return sortie


async def _analyser_groupe(
    nom_groupe: str,
    docs: list[Document],
    colonnes: list[str],
    instructions: str | None,
    model: str,
    ollama: OllamaService,
) -> dict[str, str]:
    """
    Appelle le LLM pour extraire les valeurs des colonnes pour un groupe.
    Retourne un dict {colonne: valeur}.
    """
    # Construire le contexte des documents
    parts = []
    chars_restants = 60_000
    for doc in docs:
        texte = (doc.texte_extrait or "").strip()
        if not texte:
            continue
        entete = f"\n--- {doc.nom} ---\n"
        espace = chars_restants - len(entete)
        if espace <= 0:
            break
        if len(texte) > espace:
            texte = texte[:espace] + "\n[tronqué]"
        parts.append(entete + texte)
        chars_restants -= len(entete) + len(texte)

    contexte_docs = "\n".join(parts) if parts else "(aucun document disponible)"
    colonnes_str = ", ".join(f'"{c}"' for c in colonnes)
    instructions_str = f"\nInstructions supplémentaires : {instructions}" if instructions else ""

    prompt = f"""Tu analyses les documents du candidat/société : {nom_groupe}

{contexte_docs}

Extrait les informations suivantes et retourne UNIQUEMENT un objet JSON valide, sans texte avant ni après :
{{
{chr(10).join(f'  "{c}": "valeur"' for c in colonnes)}
}}

Champs à remplir : {colonnes_str}
Si une information est absente ou non trouvée, utilise "N/A".
Ne retourne que le JSON, rien d'autre.{instructions_str}"""

    try:
        reponse = await ollama.generate(prompt, model=model)
        # Extraire le JSON de la réponse
        match = re.search(r'\{.*\}', reponse, re.DOTALL)
        if match:
            data = json.loads(match.group())
            return {c: str(data.get(c, "N/A")) for c in colonnes}
    except Exception as e:
        log.warning("Erreur parsing réponse LLM", groupe=nom_groupe, erreur=str(e))

    return {c: "N/A" for c in colonnes}


async def _run_compare(
    job_id: str,
    groupes: list[GroupeRequest],
    template_path: Path,
    colonnes: list[str],
    model: str,
    instructions: str | None,
) -> None:
    """Tâche de fond : analyse chaque groupe puis génère l'Excel."""
    from database import AsyncSessionLocal

    ollama = OllamaService()
    total = len(groupes)
    groupes_data: list[dict] = []

    try:
        async with AsyncSessionLocal() as db:
            result = await db.execute(select(Job).where(Job.id == uuid.UUID(job_id)))
            job = result.scalar_one_or_none()
            if job:
                job.statut = "running"
                job.started_at = datetime.now(tz=timezone.utc)
                await db.commit()

        for idx, groupe in enumerate(groupes, start=1):
            # Émettre "en cours"
            _compare_cache[job_id]["events"].append({
                "groupe": groupe.nom,
                "statut": "running",
                "index": idx,
                "total": total,
            })

            # Récupérer les documents
            async with AsyncSessionLocal() as db:
                doc_uuids = [uuid.UUID(did) for did in groupe.document_ids if _valid_uuid(did)]
                result = await db.execute(select(Document).where(Document.id.in_(doc_uuids)))
                docs = result.scalars().all()

            # Analyser
            valeurs = await _analyser_groupe(
                nom_groupe=groupe.nom,
                docs=list(docs),
                colonnes=colonnes,
                instructions=instructions,
                model=model,
                ollama=ollama,
            )

            groupes_data.append({"nom": groupe.nom, "valeurs": valeurs})

            # Émettre "terminé"
            _compare_cache[job_id]["events"].append({
                "groupe": groupe.nom,
                "statut": "done",
                "index": idx,
                "total": total,
            })

            log.info("Groupe analysé", job_id=job_id, groupe=groupe.nom, index=idx, total=total)

        # Générer l'Excel
        fichier = _remplir_excel(template_path, groupes_data, colonnes)
        _compare_cache[job_id]["fichier"] = str(fichier)
        _compare_cache[job_id]["statut"] = "complete"
        _compare_cache[job_id]["events"].append({
            "statut": "complete",
            "download_url": f"/api/generate/compare/download/{job_id}",
        })

        async with AsyncSessionLocal() as db:
            result = await db.execute(select(Job).where(Job.id == uuid.UUID(job_id)))
            job = result.scalar_one_or_none()
            if job:
                job.statut = "completed"
                job.completed_at = datetime.now(tz=timezone.utc)
                job.resultat = {"fichier": str(fichier), "nb_groupes": total}
                await db.commit()

        log.info("Rapport comparatif généré", job_id=job_id, fichier=str(fichier))

    except Exception as e:
        log.error("Erreur rapport comparatif", job_id=job_id, erreur=str(e))
        _compare_cache[job_id]["statut"] = "failed"
        _compare_cache[job_id]["erreur"] = str(e)
        _compare_cache[job_id]["events"].append({"statut": "failed", "erreur": str(e)})

        try:
            async with AsyncSessionLocal() as db:
                result = await db.execute(select(Job).where(Job.id == uuid.UUID(job_id)))
                job = result.scalar_one_or_none()
                if job:
                    job.statut = "failed"
                    job.erreur = str(e)
                    job.completed_at = datetime.now(tz=timezone.utc)
                    await db.commit()
        except Exception:
            pass


def _valid_uuid(val: str) -> bool:
    try:
        uuid.UUID(val)
        return True
    except ValueError:
        return False


@router.post("/generate/compare", status_code=202)
async def start_compare(
    request: CompareRequest,
    db: AsyncSession = Depends(get_db),
):
    """Lance la génération du rapport comparatif en arrière-plan."""
    # Vérifier le template
    try:
        template_uuid = uuid.UUID(request.template_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="template_id invalide")

    result = await db.execute(select(Template).where(Template.id == template_uuid))
    template = result.scalar_one_or_none()
    if not template:
        raise HTTPException(status_code=404, detail="Template non trouvé")

    template_path = Path(template.chemin_fichier)
    if not template_path.exists():
        raise HTTPException(status_code=404, detail="Fichier template introuvable sur le disque")

    colonnes = _lire_colonnes_template(template_path)
    if not colonnes:
        raise HTTPException(status_code=400, detail="Le template ne contient aucune colonne en ligne 1")

    from services import runtime_config
    model = request.model or runtime_config.model_for("rapport")

    # Créer le job
    job = Job(
        type="rapport",
        statut="pending",
        parametres={
            "type": "comparatif",
            "nb_groupes": len(request.groupes),
            "template_id": request.template_id,
            "model": model,
            "colonnes": colonnes,
        },
    )
    db.add(job)
    await db.flush()
    job_id = str(job.id)

    # Initialiser le cache
    _compare_cache[job_id] = {"events": [], "statut": "running", "fichier": None, "erreur": None}

    # Lancer en arrière-plan (asyncio.create_task — compatible avec workers=1)
    asyncio.create_task(_run_compare(
        job_id=job_id,
        groupes=request.groupes,
        template_path=template_path,
        colonnes=colonnes,
        model=model,
        instructions=request.instructions,
    ))

    log.info("Comparaison lancée", job_id=job_id, nb_groupes=len(request.groupes), colonnes=colonnes)
    return {
        "job_id": job_id,
        "statut": "en_attente",
        "nb_groupes": len(request.groupes),
        "colonnes": colonnes,
        "stream_url": f"/api/generate/compare/stream/{job_id}",
    }


@router.get("/generate/compare/stream/{job_id}")
async def stream_compare(job_id: str):
    """
    Flux SSE de progression de la comparaison.

    Événements émis :
      {"groupe": "Société A", "statut": "running", "index": 1, "total": 3}
      {"groupe": "Société A", "statut": "done",    "index": 1, "total": 3}
      {"statut": "complete", "download_url": "/api/generate/compare/download/{job_id}"}
      {"statut": "failed",   "erreur": "..."}
    """
    if not _valid_uuid(job_id):
        raise HTTPException(status_code=400, detail="ID invalide")

    async def event_generator():
        position = 0
        max_attente = 600   # 10 minutes
        attente = 0

        while attente < max_attente:
            cache = _compare_cache.get(job_id)
            if cache is None:
                yield f"data: {json.dumps({'statut': 'failed', 'erreur': 'Job introuvable'})}\n\n"
                return

            events = cache["events"]
            # Envoyer les nouveaux événements
            while position < len(events):
                yield f"data: {json.dumps(events[position])}\n\n"
                position += 1

            # Terminer si fini
            if cache["statut"] in ("complete", "failed"):
                return

            await asyncio.sleep(0.5)
            attente += 0.5

        yield f"data: {json.dumps({'statut': 'failed', 'erreur': 'Timeout'})}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@router.get("/generate/compare/download/{job_id}")
async def download_compare(job_id: str):
    """Télécharge le fichier Excel généré par la comparaison."""
    if not _valid_uuid(job_id):
        raise HTTPException(status_code=400, detail="ID invalide")

    cache = _compare_cache.get(job_id)
    if not cache:
        raise HTTPException(status_code=404, detail="Job introuvable")
    if cache["statut"] != "complete":
        raise HTTPException(status_code=400, detail="La comparaison n'est pas encore terminée")
    if not cache["fichier"]:
        raise HTTPException(status_code=500, detail="Fichier non généré")

    fichier = Path(cache["fichier"])
    if not fichier.exists():
        raise HTTPException(status_code=404, detail="Fichier introuvable sur le disque")

    return FileResponse(
        path=str(fichier),
        filename=fichier.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
